from django.http import HttpResponse
from django.shortcuts import render

# Create your views here.
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, colors
from openpyxl.worksheet import Worksheet
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.cell.cell import Cell

from cert_center.models import CertificationCenter
from digital_key.models import DigitalKey
from digital_key.support import DigitalKeyWrapper
from digital_key.views import _to_string
from employee.models import Employee, EmployeeGroup

FILLS = {
    'default': PatternFill(),
    'success': PatternFill('solid', fgColor='dff0d8'),
    'warning': PatternFill('solid', fgColor='fcf8e3'),
    'danger': PatternFill('solid', fgColor='f2dede'),
}


def get_expire_fill(key):
    wrap = DigitalKeyWrapper(key)
    style = wrap.get_label_class()
    return FILLS[style]


def export_key_xlsx(request):
    wb = Workbook()

    alignment = Alignment(horizontal='center', vertical='center')
    bold = Font(bold=True)
    fill_success = PatternFill('solid', fgColor='5CB85C')
    fill = PatternFill("solid", fgColor="DDDDDD")

    ws: Worksheet = wb.active
    ws.title = 'Ключи'
    ws.append(['Назначение', 'номер', 'тип', 'имя', 'начало', 'конец', 'фио',
               'выдан', 'группа', 'уц', 'хранение', 'описание', 'оригинал', 'системы'])
    for key in DigitalKey.objects.all().order_by('date_expire'):
        _ass = _to_string(key.assignment)
        _ser = _to_string(key.serial)
        _typ = _to_string(key.type)
        _nam = _to_string(key.name)
        _beg = key.date_begin
        _exp = Cell(ws, value=key.date_expire)
        _exp.fill = get_expire_fill(key)
        _cer = _to_string(key.cert_holder)
        _rec = _to_string(key.key_receiver)
        _emp = _to_string(key.employee_group)
        _cen = _to_string(key.cert_center)
        _loc = _to_string(key.location)
        _des = _to_string(key.description)
        _cop = _to_string(key.copy_of.serial) if key.copy_of else None
        _sys = ', '.join(sys.name for sys in key.work_systems.all())
        ws.append([_ass, _ser, _typ, _nam, _beg, _exp, _cer, _rec, _emp, _cen, _loc, _des, _cop, _sys])

    # for cols in ws.iter_cols(min_col=6, max_col=6, min_row=2):
    #     for cell in cols:
    #         cell.fill = fill_success

    ws = wb.create_sheet('Люди')
    ws.append(['Имя', 'Контакты', 'Описание'])
    for employee in Employee.objects.all():
        _nam = _to_string(employee.name)
        _con = ', '.join('{}:{}'.format(con.type.name, con.value) for con in employee.contactinfo_set.all())
        _des = _to_string(employee.description)
        ws.append([_nam, _con, _des])

    ws = wb.create_sheet('Группы')
    ws.append(['Название', 'Члены', 'Описание'])
    for group in EmployeeGroup.objects.all():
        _nam = group.name
        _mem = ', '.join(mmb.name for mmb in group.members.all())
        _des = group.description
        ws.append([_nam, _mem, _des])

    ws = wb.create_sheet('УЦ')
    ws.append(['Название', 'Ссылка', 'Описание'])
    for center in CertificationCenter.objects.all():
        _nam = center.name
        _lin = center.link
        _des = center.description
        ws.append([_nam, _lin, _des])

    # ws['A1'] = 42
    #
    # # Rows can also be appended
    # ws.append(['ключ 1', 2, 3])
    # ws.append(['ключ 1', 2, 3])
    # ws.merge_cells('A2:A3')
    # ws.append(['ключ 2', 2, 3])
    # # Python types will automatically be converted
    # import datetime
    #
    # ws['A4'] = datetime.datetime.now()

    for sheet in wb.worksheets:
        # set header format
        for cell in sheet[1]:
            cell.font = bold
            cell.alignment = alignment
        # fix column width
        for column_cells in sheet.columns:
            length = max(len(_to_string(cell.value)) for cell in column_cells)
            column_ = sheet.column_dimensions[column_cells[0].column]
            column_.width = length + 2 if length <= 20 else 22

    response = HttpResponse(save_virtual_workbook(wb), content_type=wb.mime_type)
    response['Content-Disposition'] = 'attachment; filename="keys.xlsx"'

    return response
