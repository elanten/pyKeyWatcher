from django.core.handlers.wsgi import WSGIHandler
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, get_list_or_404, redirect
from openpyxl.worksheet import Worksheet

from .forms import DigitalKeyForm
from .support import DigitalKeyWrapper
from .models import DigitalKey

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Alignment, Font
import csv


def _to_string(obj, default=''):
    return str(obj) if obj else default


def show_all(request):
    keys = DigitalKey.objects.all()
    key_views = list(DigitalKeyWrapper(key) for key in keys)
    return render(request, 'digital_key/list.html', {
        'key_views': key_views
    })


def export_key_xlsx(request):
    wb = Workbook()

    alignment = Alignment(horizontal='center', vertical='center')
    bold = Font(bold=True)

    ws: Worksheet = wb.active
    ws.title = 'Ключи'
    ws.append(['Назначение', 'номер', 'тип', 'имя', 'начало', 'конец', 'фио',
               'выдан', 'группа', 'уц', 'хранение', 'описание', 'оригинал'])

    for cell in ws[1]:
        cell.font = bold
        cell.alignment = alignment

    for key in DigitalKey.objects.all():
        _ass = _to_string(key.assignment)
        _ser = _to_string(key.serial)
        _typ = _to_string(key.type)
        _nam = _to_string(key.name)
        _beg = key.date_begin
        _exp = key.date_expire
        _cer = _to_string(key.cert_holder)
        _rec = _to_string(key.key_receiver)
        _emp = _to_string(key.employee_group)
        _cen = _to_string(key.cert_center)
        _loc = _to_string(key.location)
        _des = _to_string(key.description)
        _cop = _to_string(key.copy_of.serial) if key.copy_of else None
        ws.append([_ass, _ser, _typ, _nam, _beg, _exp, _cer, _rec, _emp, _cen, _loc, _des, _cop])

    # ws['A1'] = 42
    #
    # # Rows can also be appended
    # ws.append(['ключ 1', 2, 3])
    # ws.append(['ключ 1', 2, 3])
    # ws.merge_cells('A2:A3')
    # ws.append(['ключ 2', 2, 3])
    # # Python types will automatically be converted
    # import datetime
    #
    # ws['A4'] = datetime.datetime.now()

    # fix column width
    for column_cells in ws.columns:
        length = max(len(_to_string(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column].width = length + 2

    response = HttpResponse(save_virtual_workbook(wb), content_type=wb.mime_type)
    response['Content-Disposition'] = 'attachment; filename="keys.xlsx"'

    return response


def show_by_id(request, key_id):
    digital_key = get_object_or_404(DigitalKey, pk=key_id)
    return render(request, 'digital_key/detail.html', {
        'view': DigitalKeyWrapper(digital_key)
    })


def edit_by_id(request: WSGIHandler, key_id):
    digital_key = get_object_or_404(DigitalKey, pk=key_id)
    # holders = digital_key.contacts.filter(digitalkeycontact__type=DigitalKeyContact.HOLDER).values('id', 'name')
    # contacts = digital_key.contacts.filter(digitalkeycontact__type=DigitalKeyContact.CONTACT).values('id', 'name')

    if request.method == 'POST':
        digital_key_form = DigitalKeyForm(request.POST, instance=digital_key)
        if digital_key_form.is_valid():
            if digital_key_form.has_changed():
                digital_key = digital_key_form.save()
            digital_key.contacts.clear()
            holders_ids = request.POST.getlist('holders', [])
            contacts_ids = request.POST.getlist('contacts', [])
            # for _id in holders_ids:
            #     DigitalKeyContact(digital_key=digital_key, contragent_id=_id,
            #                       type=DigitalKeyContact.HOLDER).save()
            #
            # for _id in contacts_ids:
            #     DigitalKeyContact(digital_key=digital_key, contragent_id=_id,
            #                       type=DigitalKeyContact.CONTACT).save()
            return redirect('digital_key:show_by_id', digital_key.id)
    else:
        digital_key_form = DigitalKeyForm(instance=digital_key)

    return render(request, 'digital_key/edit.html', {
        'digitalkey_form': digital_key_form,
        # 'holders': holders,
        # 'contacts': contacts
    })


def create(request):
    digital_key = DigitalKey()
    if request.method == 'POST':
        digital_key_form = DigitalKeyForm(request.POST, instance=digital_key)
        if digital_key_form.is_valid():
            digital_key = digital_key_form.save()
            holders_ids = request.POST.getlist('holders', [])
            contacts_ids = request.POST.getlist('contacts', [])
            # for _id in holders_ids:
            #     DigitalKeyContact(digital_key=digital_key, contragent_id=_id,
            #                       type=DigitalKeyContact.HOLDER).save()
            # for _id in contacts_ids:
            #     DigitalKeyContact(digital_key=digital_key, contragent_id=_id,
            #                       type=DigitalKeyContact.CONTACT).save()
            return redirect('digital_key:show_by_id', digital_key.id)
    else:
        digital_key_form = DigitalKeyForm(instance=digital_key)
    return render(request, 'digital_key/edit.html', {
        'digitalkey_form': digital_key_form,
        'holders': [],
        'contacts': []
    })


def remove_by_id(request, key_id):
    digital_key = get_object_or_404(DigitalKey, pk=key_id)
    digital_key.delete()
    return redirect('digital_key:all')
