from openpyxl import load_workbook

from digital_key.models import *
from employee.models import *
from cert_center.models import *
from key_manual.models import *

from django.utils import dateparse, timezone

file_path = 'exclude/base.xlsx'

wb = load_workbook(file_path)
ws = wb.active

list_assignments = set(c.value for c in ws['A'][1:] if c.value)
list_types = set(c.value for c in ws['C'][1:] if c.value)
list_employees = set(c.value for c in ws['G'][1:] if c.value) \
    .union(set(c.value for c in ws['H'][1:] if c.value))
list_groups = set(c.value for c in ws['I'][1:] if c.value)
list_centers = set(c.value for c in ws['J'][1:] if c.value)
list_locations = set(c.value for c in ws['K'][1:] if c.value)

assignments = {}
types = {}
employees = {}
locations = {}
groups = {}
centers = {}
keys = {}

KeyAssignment.objects.all().delete()
for name in list_assignments:
    assignments[name] = KeyAssignment.objects.create(name=name)

KeyType.objects.all().delete()
for name in list_types:
    types[name] = KeyType.objects.create(name=name)

KeyLocation.objects.all().delete()
for name in list_locations:
    locations[name] = KeyLocation.objects.create(name=name)

Employee.objects.all().delete()
for name in list_employees:
    employees[name] = Employee.objects.create(name=name)

EmployeeGroup.objects.all().delete()
for name in list_groups:
    groups[name] = EmployeeGroup.objects.create(name=name)

CertificationCenter.objects.all().delete()
for name in list_centers:
    centers[name] = CertificationCenter.objects.create(name=name)

DigitalKey.objects.all().delete()
for row in ws.iter_rows('A2:M39'):
    assign, number, typ, key, start, end, cert, hold, group, center, loc, desc, copy = map(lambda x: x.value, row)
    keys[number] = DigitalKey.objects.create(
        name=(key or 'NoName'), serial=number, type=types.get(typ, None), description=(desc or ''),
        assignment=assignments.get(assign, None), location=locations.get(loc, None),
        cert_holder=employees.get(cert, None), key_receiver=employees.get(hold, None),
        date_begin=start, date_expire=end, copy_of=(keys[copy] if copy else None),
        cert_center=(centers[center] if center else None)
    )

wb.close()
